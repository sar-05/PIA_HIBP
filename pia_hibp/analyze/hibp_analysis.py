from statistics import mean
from statistics import mode
from statistics import median
from statistics import stdev
import re
from openpyxl import Workbook
import matplotlib.pyplot as plt

from pia_hibp import config


def turn_to_dict(texto):
    result = {}
    match = re.search(r'Counter\(\{(.+)\}\)', texto)
    if match:
        data = match.group(1)
    else:
        data = texto.strip().strip('{}')
    pairs = data.split(",")
    for pair in pairs:
        if ':' in pair:
            key, value = pair.split(":", 1)
            key = key.strip().strip("'\"")
            value = value.strip().strip("})")
            result[key] = int(value)
    return result

def turn_to_list_tuples(data):
    result = []
    data = data.strip().strip("[]")
    elements = data.split("),")
    for element in elements:
        element = element.strip().strip("()").replace("'", "")
        if ',' in element:
            year, times = element.split(",", 1)
            result.append((year.strip(), int(times.strip())))
    return result

def rearrange_by_value(tupla):
    return tupla[1]

def main_hibp_analyisis():
    with open(config.HIBP_TXT, "r") as file:
        lines = file.readlines()
    count_dict = {}
    sorted_years = []
    data_classes = {}
    for i in range(len(lines)):
        line = lines[i]
        if "Breaches by name" in line:
            count_dict = turn_to_dict(lines[i+1])
        elif "Breaches by date" in line:
            sorted_years = turn_to_list_tuples(lines[i+1])
        elif "Data types leaked" in line:
            data_classes = turn_to_dict(lines[i+1])

    sorted_by_freq = sorted(sorted_years, key=rearrange_by_value, reverse=True)
    top_3_years = sorted_by_freq[:3]
    bottom_3_years = sorted_by_freq[-3:]

    sorted_pages = sorted(count_dict.items(), key=rearrange_by_value, reverse=True)
    top_pages = sorted_pages[:5]
    bottom_pages = sorted_pages[-5:]

    # Estadísticas
    valores = list(count_dict.values())
    promedio = mean(valores)
    mediana = median(valores)
    moda = mode(valores)
    desviacion = stdev(valores)

    # Tipo de dato más vulnerable
    datos_vulnerables = sorted(data_classes.items(), key=rearrange_by_value, reverse=True)
    top_data_type = datos_vulnerables[:5]

    #Excel Workbook
    book = Workbook()

    sheet1 = book.active
    sheet1.title = "Top PwnedCount websites"
    sheet1["A1"] = "Websites with more breaches"
    sheet1["B1"] = "PwnedCount"
    for page, times in top_pages:
        sheet1.append([page, times])

    sheet2 = book.create_sheet("Bottom PwnedCount websites")
    sheet2["A1"] = "Websites with less breaches"
    sheet2["B1"] = "PwnedCount"
    for page, times in bottom_pages:
        sheet2.append([page, times])

    sheet3 = book.create_sheet("Years with more breaches")
    sheet3["A1"] = "Years with more breaches"
    sheet3["B1"] = "Total Breaches"
    for year, times in top_3_years:
        sheet3.append([year, times])

    sheet4 = book.create_sheet("Years with less breaches")
    sheet4["A1"] = "Years with less breaches"
    sheet4["B1"] = "Total Breaches"
    for year, times in bottom_3_years:
        sheet4.append([year, times])

    sheet5 = book.create_sheet("Top data classes")
    sheet5["A1"] = "Data classes in more breaches"
    sheet5["B1"] = "Times repeated"
    for data, times in top_data_type:
        sheet5.append([data, times])
    
    sheet6 = book.create_sheet("Statistics of accounts beached in a breached website")
    sheet6["A1"] = "Average"
    sheet6["A2"] = "Meadian"
    sheet6["A3"] = "Mode"
    sheet6["A4"] = "Standard deviation"
    sheet6["B1"] = f"{promedio:.2f}"
    sheet6["B2"] = f"{mediana:.2f}"
    sheet6["B3"] = f"{moda:}"
    sheet6["B4"] = f"{desviacion:.2f}"

    book.save(f"{config.HIBP_DIR}/analysis_HIBP.xlsx")

    #Graphs
    breaches_y = [int(times) for _, times in sorted_years]
    years = range(2007, 2026)
    fig, ax = plt.subplots()
    ax.plot(years, breaches_y, marker = "*")
    ax.set_title("Breaches by year")
    ax.set_xlabel("Year")
    ax.set_ylabel("Amount of breaches")
    ax.axis(ymin=0, ymax=103, xmin=2007, xmax=2025)
    plt.savefig(f"{config.HIBP_DIR}/breaches_by_year.png")
    plt.close(fig)

    top_labels = [page for page, _ in top_pages]
    top_values = [count for _, count in top_pages]
    fig, ax = plt.subplots()
    ax.bar(top_labels, top_values, color='tomato')
    ax.set_title("Top 5 breached websites")
    ax.set_xlabel("Websites")
    ax.set_ylabel("Accounts breached")
    plt.savefig(f"{config.HIBP_DIR}/most_breached.png")
    plt.close(fig)

    bottom_labels = [page for page, _ in bottom_pages]
    bottom_values = [count for _, count in bottom_pages]
    fig, ax = plt.subplots()
    ax.bar(bottom_labels, bottom_values, color='aqua')
    ax.set_title("Bottom 5 breached websites")
    ax.set_xlabel("Websites")
    ax.set_ylabel("Accounts breached")
    plt.savefig(f"{config.HIBP_DIR}/least_breached.png")
    plt.close(fig)

    data_types = [x[0] for x in top_data_type]
    v_data_type = [x[1] for x in top_data_type]
    fig, ax = plt.subplots()
    ax.pie(v_data_type, labels=data_types)
    ax.set_title("Top 5 data classes in breaches")
    plt.savefig(f"{config.HIBP_DIR}/top_data_classes.png")
    plt.close(fig)   

    # Printing results
    print("Tres años con más filtraciones:")
    for year, times in top_3_years:
        print(f"Año: {year} , {times} filtraciones")

    print("\nTres años con menos filtraciones:")
    for year, times in bottom_3_years:
        print(f"Año: {year} , {times} filtraciones")

    print("\nSitios con más cuentas comprometidas:")
    for page, times in top_pages:
        print(f"{page} : {times:,} cuentas")

    print("\nSitios con menos cuentas comprometidas:")
    for page, times in bottom_pages:
        print(f"{page} : {times:,} cuentas")

    print(f"\nPromedio de cuentas comprometidas por sitio: {promedio:.2f}")
    print(f"Mediana: {mediana:.2f}")
    print(f"Moda: {moda:}")
    print(f"Desviación estándar: {desviacion:.2f}")

    print(f"\nTop tipos de datos más vulnerable:")
    for data, times in top_data_type:
        print(f"tipo de dato: {data}, filtraciones: {times}")
